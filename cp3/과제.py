import requests
from bs4 import BeautifulSoup
import openpyxl

# 엑셀 준비
wb = openpyxl.Workbook()
ws = wb.create_sheet('주식현재가')

ws['A1'] = '종목'
ws['B1'] = '현재가'

ws['A2'] = '삼성전자'
ws['A3'] = 'sk하이닉스'
ws['A4'] = '카카오'

#종목 코드 리스트
codes = [
  '005930',
  '000660',
  '035720'
]
prices = []
for i in codes:
  url = f"https://finance.naver.com/item/sise.naver?code={i}"
  
  response = requests.get(url)
  
  html = response.text
  
  soup = BeautifulSoup(html, 'html.parser')
  
  #현재가 텍스트만 변수에 저장
  price = soup.select_one("#_nowVal").text
  
  #변수의 타입 변경 용이하게 하기 위해 ','를''로 바꿔줌 
  price = price.replace(',','')
  
  prices.append(price)
  
ws['B2'] = prices[0]
ws['B3'] = prices[1]
ws['B4'] = prices[2]

wb.save(r'주식현재가.xlsx')