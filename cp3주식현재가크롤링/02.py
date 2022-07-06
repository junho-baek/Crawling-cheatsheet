import openpyxl

fpath = '학생정보.xlsx'

# 1) 엑셀 불러오기
wb = openpyxl.load_workbook(fpath)


# 2) 엑셀 시트 선택
ws = wb['학생 정보']

# 3) 데이터 수정하기
ws['A3'] = 2020122222
ws['B3'] ='junho-baek'

# 4) 엑셀 저장하기

wb.save(fpath)