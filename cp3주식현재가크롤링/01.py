import openpyxl

# 1) 엑셀 만들기
wb = openpyxl.Workbook()

# 2) 엑셀 워크시트 만들기

ws = wb.create_sheet('학생 정보')

# 3) 데이터 추가하기

ws['A1'] = '학번'
ws['B1'] = '이름'

ws['A2'] = 2020123456
ws['B2'] = '백준호'

# 4) 엑셀 저장하기

wb.save(r'학생정보.xlsx')